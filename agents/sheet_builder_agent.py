import asyncio
import json
import logging
import os
import threading
import time
import random
from collections import defaultdict
from pathlib import Path

import nats
import openpyxl
import psycopg
from dotenv import load_dotenv

# Cross-platform file locking
try:
    import portalocker
    def acquire_lock(f):
        portalocker.lock(f, portalocker.LOCK_EX)
    def release_lock(f):
        portalocker.unlock(f)
except ImportError:
    _file_lock = threading.Lock()
    def acquire_lock(f):
        _file_lock.acquire()
    def release_lock(f):
        _file_lock.release()

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [sheet_builder] %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler("logs/sheet_builder.log"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

NATS_URL   = os.getenv("NATS_URL", "nats://localhost:4222")
DSN        = os.getenv("POSTGRES_DSN")
OUTPUTS    = Path(os.getenv("OUTPUTS_DIR", "outputs"))
STATE_FILE = "data/system_state.json"

SHEET_HEADERS = [
    "retailer_id", "transaction_date", "secondary_transaction_id",
    "product_category_snapshot", "sku_name",
    "secondary_gross_value", "secondary_tax_amount", "secondary_net_value",
]

# In-memory buffer: distributor_id → list of row dicts
# This accumulates during a session and is used to rebuild sheets
buffer: dict[str, list[dict]] = defaultdict(list)


def group_by_retailer(rows: list[dict]) -> dict[str, list[dict]]:
    out = defaultdict(list)
    for r in sorted(rows, key=lambda x: (x["retailer_id"], x["transaction_date"])):
        out[r["retailer_id"]].append(r)
    return dict(out)


def write_distributor_sheet(dist_id: str, rows: list[dict]) -> None:
    """Write all transactions for a distributor into one Excel workbook."""
    OUTPUTS.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    grouped = group_by_retailer(rows)

    for retailer_id, txns in grouped.items():
        sheet_name = retailer_id[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_name)

        # Headers
        ws.append(SHEET_HEADERS)
        # Style header row bold
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for t in txns:
            ws.append([t.get(col) for col in SHEET_HEADERS])

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    out_path = OUTPUTS / f"{dist_id}.xlsx"
    wb.save(str(out_path))
    log.info(f"Written: {out_path} ({len(rows)} rows, {len(grouped)} retailers)")


def update_system_state(row: dict, max_retries: int = 5) -> bool:
    """Add new transaction to system_state.json with file locking retry."""
    txn_id = str(row.get("secondary_transaction_id", ""))
    if not txn_id:
        log.warning("No transaction ID found in row, skipping state update")
        return False

    # Ensure data directory exists
    Path(STATE_FILE).parent.mkdir(parents=True, exist_ok=True)

    for attempt in range(max_retries):
        try:
            # Load existing state or create new
            state = {}
            if os.path.exists(STATE_FILE):
                with open(STATE_FILE, "r") as f:
                    acquire_lock(f)
                    try:
                        content = f.read()
                        state = json.loads(content) if content else {}
                    finally:
                        release_lock(f)

            # Skip if transaction already exists
            if txn_id in state:
                log.debug(f"Transaction {txn_id} already in state file")
                return True

            # Add new transaction with initial state
            state[txn_id] = {
                "distributor_id": row.get("distributor_id", ""),
                "retailer_id": row.get("retailer_id", ""),
                "sku_name": row.get("sku_name", ""),
                "transaction_date": row.get("transaction_date", ""),
                "gross_value": row.get("secondary_gross_value", 0),
                "tax_amount": row.get("secondary_tax_amount", 0),
                "net_value": row.get("secondary_net_value", 0),
                "mail_status": False,
                "reply_status": False,
                "reply_content": None
            }

            # Write back atomically
            temp_file = STATE_FILE + ".tmp"
            with open(temp_file, "w") as f:
                json.dump(state, f, indent=2)
            os.replace(temp_file, STATE_FILE)

            log.info(f"Added transaction {txn_id} to system_state.json")
            return True

        except (IOError, OSError, PermissionError) as e:
            if attempt < max_retries - 1:
                wait_time = (2 ** attempt) + random.uniform(0, 0.1)
                log.warning(f"File locked, retrying in {wait_time:.2f}s... (attempt {attempt + 1}/{max_retries})")
                time.sleep(wait_time)
            else:
                log.error(f"Failed to update system_state.json for txn {txn_id} after {max_retries} attempts: {e}")
                return False
        except Exception as e:
            log.error(f"Unexpected error updating system_state.json for txn {txn_id}: {e}")
            return False

    return False


async def write_to_db(row: dict, db) -> None:
    await db.execute("""
        INSERT INTO transactions (
            distributor_id, retailer_id, sku_name,
            product_category_snapshot, secondary_transaction_id,
            transaction_date, secondary_gross_value,
            secondary_tax_amount, secondary_net_value
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (secondary_transaction_id) DO NOTHING
    """, (
        row["distributor_id"], row["retailer_id"], row["sku_name"],
        row["product_category_snapshot"], row["secondary_transaction_id"],
        row["transaction_date"], row["secondary_gross_value"],
        row["secondary_tax_amount"], row["secondary_net_value"],
    ))
    await db.commit()


async def main():
    nc  = await nats.connect(NATS_URL)
    sub = await nc.subscribe("transaction.ingested")
    log.info("Sheet builder agent running — subscribed to transaction.ingested")

    async with await psycopg.AsyncConnection.connect(DSN) as db:
        async for msg in sub.messages:
            try:
                row = json.loads(msg.data)
                dist_id = row["distributor_id"]

                # Write to DB
                await write_to_db(row, db)

                # Add to system_state.json for dashboard visibility
                update_system_state(row)

                # Add to in-memory buffer and rebuild sheet
                buffer[dist_id].append(row)
                write_distributor_sheet(dist_id, buffer[dist_id])

            except Exception as e:
                log.error(f"Error processing message: {e}", exc_info=True)


if __name__ == "__main__":
    import selectors
    loop_factory = lambda: asyncio.SelectorEventLoop(selectors.SelectSelector())
    asyncio.run(main(), loop_factory=loop_factory)
